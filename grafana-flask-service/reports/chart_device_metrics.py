from datetime import datetime
from reports.separate_report_xls import SeparateReportXLs
from sciencelogic_db import ScienceLogicDb
from reports.report_utils import ReportUtility
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

wb = Workbook()
ws = wb.active

utils = ReportUtility()
xls = SeparateReportXLs()


class ChartDeviceMetrics:
    def __init__(self):
        self.sciencelogic_db = ScienceLogicDb()

    def get_report(self, from_date, to_date, organizations, devices_by_organization):
        result_dict = self.generate_report(
            from_date, to_date, organizations, devices_by_organization)
        return self.generate_chart(
            result_dict, from_date, to_date, devices_by_organization)

    def generate_report(
        self, from_date, to_date, organizations, devices_by_organization
    ):
        query = utils.frame_query("chart_device_utilization_metrics")
        params = (
            from_date, to_date, organizations, devices_by_organization,
            from_date, to_date, organizations, devices_by_organization,
            from_date, to_date, organizations, devices_by_organization,)
        final_dict = {}
        super_final_dict = {}
        device_dict = {}
        rows = self.sciencelogic_db.execute(query, params, True)
        rows.pop(0)
        for row in rows:
            device_name = row[0]
            if device_dict.get(device_name) is None:
                row.pop(0)
                temp_list = [row]
                device_dict[device_name] = temp_list
            else:
                row.pop(0)
                device_dict.get(device_name).append(row)

        for key, value in device_dict.items():
            dt_dict = {}
            for v in value:
                dt = v[0]
                if dt_dict.get(dt) is None:
                    v.pop(0)
                    temp_list = [v]
                    dt_dict[dt] = temp_list
                else:
                    v.pop(0)
                    dt_dict.get(dt).append(v)
            final_dict[key] = dt_dict

        for key, value in final_dict.items():
            super_final_dict[key] = []
            for ki, val in value.items():
                temp_list = [ki]
                if len(val) == 3:
                    for v in val:
                        for j in v:
                            if j is not None and len(str(j)) > 0:
                                temp_list.append(float(j))
                elif len(val) == 2:
                    dupli_list = [1, 2, 3]
                    for v in val:
                        count = 1
                        for j in v:
                            if j is not None and len(str(j)) > 0:
                                temp_list.append(float(j))
                                break
                            count += 1
                        dupli_list.remove(count)
                    temp_list.insert(dupli_list[0], 0)
                elif len(val) == 1:
                    for v in val:
                        for j in v:
                            if j is not None and len(str(j)) > 0:
                                temp_list.append(float(j))
                            else:
                                temp_list.append(0)

                super_final_dict.get(key).append(temp_list)
        return super_final_dict

    def generate_chart(self, result_dict, from_date, to_date, devices_by_organization):

        wb = Workbook()
        ws = wb.active
        we = wb.create_sheet("data", 1)
        we.sheet_state = 'hidden'
        img = openpyxl.drawing.image.Image(
            utils.get_poly_logo() + 'poly_report_logo.jpg')
        img.width = 380
        img.height = 65
        ws.add_image(img, 'A1')
        chart_row = 7
        fontStyle = Font(size="16", bold=True)
        ws['A' + str(chart_row)] = "Device Utilization Chart " + str(from_date) + " to " + str(
            to_date) + str(" - Based on UTC timezone")
        ws['A' + str(chart_row)].font = fontStyle
        ws.merge_cells('A' + str(chart_row) + ':P' + str(chart_row))
        chart_row += 1
        # writing geretated date time
        ws['N1'] = "Generated on: " + str(utils.getDateTime())
        ws['N2'] = "Begin: : " + str(from_date)
        ws['N3'] = "End: " + str(to_date)
        chart_row += 2
        previous_row = 0
        for title, results in result_dict.items():
            self.write_data_on_sheet(ws, chart_row, "Device: "+str(title))
            chart_row += 2
            headers = ['dates', "CPU Usage", "RAM Usage", "SWAP Usage"]
            results.insert(0, headers)
            from_flag = True
            to_flag = True
            count = 0
            for row in results:
                if from_date in row:
                    from_flag = False
                if to_date in row:
                    to_flag = False
                we.append(row)
            if from_flag:
                we.insert_rows(previous_row+2, amount=1)
                we.cell(previous_row+2, 1, from_date)
                count += 1
            if to_flag:
                we.append([to_date])
                count += 1
            from datetime import datetime
            if len(results) > 1:
                c2 = LineChart()
                c2.title = str(title)
                # c2.style = 12
                c2.y_axis.title = "Percentage"
                c2.y_axis.crossAx = 500
                c2.height = 10  # default is 7.5
                width = len(results) - 1
                # width = width if width > 15 else 15
                width = 30
                c2.width = width
                c2.x_axis = DateAxis(crossAx=100)
                c2.x_axis.number_format = 'd-mmm-yyyy HH:MM:SS'
                c2.x_axis.majorTimeUnit = "days"
                c2.x_axis.title = "Date"
                max_col = len(results[0])
                data = Reference(
                    we, min_col=2, min_row=previous_row+1,
                    max_col=max_col, max_row=previous_row+len(results)+count)
                c2.add_data(data, titles_from_data=True)
                dates = Reference(
                    we, min_col=1, min_row=previous_row+2,
                    max_row=previous_row+len(results)+count)
                c2.series[1].graphicalProperties.line.solidFill = "ff4210"
                c2.series[0].graphicalProperties.line.solidFill = "0b407f"
                c2.series[2].graphicalProperties.line.solidFill = "f4d53f"
                # c2.series[0].graphicalProperties.line.width = 20000
                # c2.series[2].graphicalProperties.line.width = 20000
                # c2.series[1].graphicalProperties.line.width = 20000

                flag = True
                if len(results) == 2:
                    flag = False
                    for c in c2.series:
                        # 12700 emu = 1 pt(72*1 inches)
                        c.graphicalProperties.line.width = 20000
                        c.marker.symbol = 'diamond'
                if flag:
                    for c in c2.series:
                        # 12700 emu = 1 pt(72*1 inches)
                        c.graphicalProperties.line.width = 20000

                # for c in c2.series:
                #     c.graphicalProperties.line.width = 20000    # 12700 emu = 1 pt(72*1 inches)

                c2.set_categories(dates)
                ws.add_chart(c2, "A"+str(chart_row))
                previous_row += len(results) + count
                chart_row += 25
            else:
                ws['A' + str(chart_row)
                   ] = "No Data Found for selected device â€“ "
                ws['A' + str(chart_row)].font = fontStyle
                chart_row += 25
        import datetime
        x = datetime.datetime.now()
        date_time_str = x.strftime("%d_%m_%YT%H_%M_%S")
        file_name = "xls_data_" + utils.gen_random() + str(date_time_str + ".xlsx")
        wb.save(str(utils.get_report_path()) + file_name)
        return file_name

    def write_data_on_sheet(self, ws, chart_row, data):
        font = Font(bold=True, size=20)
        ws['A' + str(chart_row)] = data
        ws['A' + str(chart_row)].font = font
        ws['A' + str(chart_row)].fill = PatternFill(
            start_color="00999999",
            end_color="999999", fill_type="solid")
        ws.merge_cells('A' + str(chart_row) + ':P' + str(chart_row))
