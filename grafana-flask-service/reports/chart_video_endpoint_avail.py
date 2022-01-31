from reports.separate_report_xls import SeparateReportXLs
from sciencelogic_db import ScienceLogicDb
from reports.report_utils import ReportUtility
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import openpyxl
import datetime
from openpyxl.chart import (
    Reference,
    Series,
    BarChart3D,
)

utils = ReportUtility()
xls = SeparateReportXLs()


class ChartVideoEndpointAvail:
    def __init__(self):
        self.sciencelogic_db = ScienceLogicDb()


    def get_report(self, from_date, to_date, organizations, devices_by_organization):
        results = self.generate_report(from_date, to_date, organizations, devices_by_organization)
        return self.generate_chart(results, from_date, to_date)

    def generate_report(self, from_date, to_date, organizations, devices_by_organization):
        query = utils.frame_query("video_endpoint_availlability")
        params = (organizations, devices_by_organization, from_date, to_date,)
        results = self.sciencelogic_db.execute(query, params, True)
        results.pop(0)
        return results

    def generate_chart(self, rows, from_date, to_date):
        wb = Workbook()
        ws = wb.active
        we = wb.create_sheet("data", 1)
        we.sheet_state = 'hidden'
        img = openpyxl.drawing.image.Image(utils.get_poly_logo() + 'poly_report_logo.jpg')
        img.width = 380
        img.height = 65
        ws.add_image(img, 'A1')
        chart_row = 7
        max_row = 2
        min_row = 2
        fontStyle = Font(size="16", bold=True)
        ws['A' + str(chart_row)] = "Video Endpoint Availability Chart (percent) – " + str(from_date) + " to " + str(
            to_date)
        ws['A' + str(chart_row)].font = fontStyle
        ws.merge_cells('A' + str(chart_row) + ':M' + str(chart_row))

        # writing geretated date time
        ws['N1'] = "Generated on: " + str(utils.getDateTime())
        ws['N2'] = "Begin: : " + str(from_date)
        ws['N3'] = "End: " + str(to_date)

        chart_row += 2

        for row in rows:
            # writing data on sheet
            org = "Organization: "+str(row[0])
            self.write_data_on_sheet(ws, chart_row, org)
            chart_row += 1
            row.pop(0)
            # device_type = "Device Group: "+str(row[0])
            # self.write_data_on_sheet(ws, chart_row, device_type)
            # chart_row += 1
            row.pop(0)
            device_category = "Category: "+str(row[0]).replace("."," ")
            self.write_data_on_sheet(ws, chart_row, device_category)
            chart_row += 1
            row.pop(0)
            
            report_device = "Device: " + str(row[0])
            point_value = row[len(row)-1]
            row[len(row)-1] = round(float(point_value),2)
            we.append([None, 'value'])
            we.append(row)
            chart_row += 1
            data = Reference(we, min_col=2, min_row=min_row, max_col=2, max_row=max_row)
            titles = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            min_row += 2
            max_row += 2
            chart = BarChart3D()
            chart.y_axis.title = 'Percentage'
            chart.title = report_device
            chart.add_data(data=data, titles_from_data=True)
            chart.set_categories(titles)
            ws.add_chart(chart, "A" + str(chart_row))
            chart_row += 20
        x = datetime.datetime.now()
        date_time_str = x.strftime("%d_%m_%YT%H_%M_%S")
        file_name = "xls_data_" + utils.gen_random() + str(date_time_str + ".xlsx")
        wb.save(str(utils.get_report_path()) + file_name)
        return file_name

    def write_data_on_sheet(self, ws, chart_row, data):
        font = Font(bold=True)
        ws['A' + str(chart_row)] = data
        ws['A' + str(chart_row)].font = font
        ws['A' + str(chart_row)].fill = PatternFill(start_color="00999999", end_color="999999", fill_type="solid")
        ws.merge_cells('A' + str(chart_row) + ':F' + str(chart_row))
