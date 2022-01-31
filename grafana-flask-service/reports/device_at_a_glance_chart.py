import openpyxl
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    BarChart3D,
    PieChart,
    PieChart3D,
    LineChart,

)
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Font, Font
from openpyxl.styles import PatternFill
from reports.report_utils import ReportUtility
from datetime import datetime
import datetime
from datetime import date


utils = ReportUtility()

class DeviceAtAGlanceChart:
    def write_data_xls(self, result_dict, from_date, to_date):
        try:
            wb = Workbook()
            ws = wb.active
            we = wb.create_sheet("data", 1)
            we2 = wb.create_sheet("data2", 1)
            we3 = wb.create_sheet("data3", 1)
            we.sheet_state = 'hidden'
            we2.sheet_state = 'hidden'
            we3.sheet_state = 'hidden'
            img = openpyxl.drawing.image.Image(utils.get_poly_logo() + 'poly_report_logo.jpg')
            img.width = 380
            img.height = 65
            ws.add_image(img, 'A1')
            chart_row = 7
            max_row = 2
            min_row = 2
            fontStyle = Font(size="16", bold=True)
            ws['A' + str(chart_row)] = "Device At A Glance " + str(from_date) + " to " + str(
                to_date)
            ws['A' + str(chart_row)].font = fontStyle
            ws.merge_cells('A' + str(chart_row) + ':P' + str(chart_row))
            chart_row +=1
            ws['N1'] = "Generated on: " + str(utils.getDateTime())
            ws['N2'] = "Begin: : " + str(from_date)
            ws['N3'] = "End: " + str(to_date)
            chart_row += 2
            previous_row = 0

            data_row = 1
            ava_row = 1
            schedule_row = 1
            chart_col= 1
            print(type(result_dict))
            for title in result_dict.keys():
                font = Font(bold=True, size=16)
                report_device = "Device: "+ str(title)
                ws['A' + str(chart_row)] = report_device
                ws['A' + str(chart_row)].font = font
                ws['A' + str(chart_row)].fill = PatternFill(start_color="00999999", end_color="999999", fill_type="solid")
                ws.merge_cells('A' + str(chart_row) + ':P' + str(chart_row))

                # headers = ['device Name', 'id', 'CPU MIN', 'CPU AVG', 'CPU Max', 'MEM MIN', 'MEM AVG', 'MEM MAX', 'SWAP MIN', 'SWAP AVG', 'SWAP MAX', 'Avail', 'Schedule', 'unschedule']
                # result_dict.insert(0, headers)
                # for row in result:
                #     we.append(row)
                results = result_dict[title]
                if len(results) > 0:
                    cpu_lis = ['CPU % Used']
                    mem_lis = ['Memory % Used']
                    swap_lis = ['Swap % Used']
                    # if result[0] or result[1] or result[2]:
                    chart_row += 1
                    if results[0]:
                        cpu_lis.extend(results[0])
                    if results[1]:
                        mem_lis.extend(results[1])
                    if results[2]:
                        swap_lis.extend(results[2])
                    we.append(['Type', 'Min', 'Avg', 'Max'])
                    we.append(cpu_lis)
                    we.append(mem_lis)
                    we.append(swap_lis)
                    we.append([])
                    chart1 = BarChart3D()
                    chart1.y_axis.title = 'Percentage'
                    chart1.title = 'System Vitals'
                    # chart1.y_axis.crossAx = 500
                    # chart1.height = 10  # default is 7.5
                    # width = len(result) - 1
                    # width = width if width > 15 else 15
                    # c2.width = width

                    data = Reference(we, min_col=2, min_row=data_row, max_col=4, max_row=data_row+3)
                    titles = Reference(we, min_col=1, min_row=data_row+1, max_row=data_row+3)
                    data_row +=5
                    # print(data_row)
                    chart1.add_data(data=data, titles_from_data=True)
                    chart1.set_categories(titles)

                    ws.add_chart(chart1, "A"+str(chart_row))
                    chart_row +=18

                    if results[3]:
                        chart_row += 2
                        d = results[3]
                        we2.append(['Type', 'Stats'])
                        we2.append(['Available', d[0]])
                        we2.append(['Unavailable', d[1]])
                        we2.append([])
                        chart2 = PieChart()
                        data = Reference(we2, min_col=2, min_row=ava_row, max_col=3, max_row=ava_row+2)
                        titles = Reference(we2, min_col=1, min_row=ava_row + 1, max_row=ava_row+2)
                        chart2.add_data(data=data, titles_from_data=True)
                        chart2.set_categories(titles)
                        ava_row +=4
                        chart2.title = "Availability"
                        # chart2.height =  11
                        chart2.width = 10
                        # slice = DataPoint(idx=0, explosion=20)
                        # chart2.series[0].data_points = [slice]
                        ws.add_chart(chart2, "A"+str(chart_row))
                        # ws.add_chart(data=data, titles_from_data=True)
                        chart_row +=18


                    scd_uns_dic = {}
                    temp_scd = results[4]
                    temp_uns = results[5]
                    for val in temp_scd:
                        if scd_uns_dic.get(val[0]) is None:
                            scd_uns_dic[val[0]] = [int(val[1])]
                        else:
                            scd_uns_dic[val[0]].append(int(val[1]))
                    for val in temp_uns:
                        if scd_uns_dic.get(val[0]) is None:
                            scd_uns_dic[val[0]] = [0, int(val[1])]
                        else:
                            scd_uns_dic[val[0]].append(int(val[1]))
                    # print(scd_uns_dic)
                    if temp_scd or temp_uns:
                        chart_row +=2
                        we3.append(['', 'Scheduled', 'Unscheduled'])
                        count_row = 0
                        for key in scd_uns_dic.keys():
                            # print()
                            month_year = datetime.datetime.strptime(key, "%Y-%m")
                            values = [month_year]
                            values.extend(scd_uns_dic[key])
                            # print(values)
                            if len(values) == 2:
                                values.extend(0)
                            # print("Val", values)
                            we3.append(values)
                            count_row += 1

                        # we3.append([])
                        chart3 = LineChart()
                        chart3.x_axis.number_format = 'mmm-yyyy'
                        # c2.x_axis.majorTimeUnit = "days"
                        chart3.x_axis.title = "Date"
                        chart3.x_axis.majorTimeUnit = "Month"
                        chart3.y_axis.title = 'Outage Count'
                        chart3.title = "Maintenance Trend"

                        data = Reference(we3, min_col=2, min_row=schedule_row, max_col=3, max_row=schedule_row+count_row)
                        title_chart3 = Reference(we3, min_col=1, min_row=schedule_row+1, max_row=schedule_row+count_row)
                        chart3.add_data(data, titles_from_data= True)
                        chart3.set_categories(title_chart3)

                        line = chart3.series[0]
                        line.marker.symbol = "diamond"
                        line2 = chart3.series[1]
                        line2.marker.symbol = "x"
                        for c in chart3.series:
                            c.graphicalProperties.line.width = 20000

                        ws.add_chart(chart3, "A" + str(chart_row))
                        schedule_row += count_row+1
                        chart_row += 18

            x = datetime.datetime.now()
            date_time_str = x.strftime("%d_%m_%YT%H_%M_%S")
            file_name = "xls_data_" + utils.gen_random() + str(date_time_str + ".xlsx")
            wb.save(str(utils.get_report_path()) + file_name)
            # print(file_name)
            return file_name
        except Exception as ex:
            print("Exception while writing xls file:  " + str(ex))
            return "Exception while writing xls file:  " + str(ex), "error"

