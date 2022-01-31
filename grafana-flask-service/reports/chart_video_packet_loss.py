from datetime import datetime
from reports.separate_report_xls import SeparateReportXLs
from reports.report_utils import ReportUtility
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

from sciencelogic_db import ScienceLogicDb

wb = Workbook()
ws = wb.active

utils = ReportUtility()
xls = SeparateReportXLs()


class ChartVideoEndPointPacketLoss:
    def __init__(self):
        self.sciencelogic_db = ScienceLogicDb()

    def get_report(self, from_date, to_date, organizations, devices_by_organization):
        wb = Workbook()
        ws = wb.active
        img = openpyxl.drawing.image.Image(
            utils.get_poly_logo() + 'poly_report_logo.jpg')
        img.width = 380
        img.height = 65
        ws.add_image(img, 'A1')
        chart_row = 7
        fontStyle = Font(size="16", bold=True)
        ws['A' + str(chart_row)] = "Video Endpoint Packet Loss – " \
            + str(from_date) + " to " + str(
            to_date) + str(" - Based on UTC timezone")
        ws['A' + str(chart_row)].font = fontStyle
        ws.merge_cells('A' + str(chart_row) + ':O' + str(chart_row))
        chart_row += 1

        # writing geretated date time
        ws['N1'] = "Generated on: " + str(utils.getDateTime())
        ws['N2'] = "Begin: : " + str(from_date)
        ws['N3'] = "End: " + str(to_date)

        query = utils.frame_query('chart_packet_loss')
        params = (organizations, devices_by_organization,)
        result_devices = self.sciencelogic_db.execute(query, params, True)
        result_devices.pop(0)
        min_row = 0
        if len(result_devices) > 0:

            for device in result_devices:
                device_name = str(device).replace(
                    "'", "").replace("[", "").replace("]", "")
                chart_row += 4
                txt_query_list = [
                    'packet_loss_content',
                    'packet_loss_video',
                    'packet_loss_audio']
                self.write_data_on_sheet(
                    ws, chart_row, "Device: " + device_name)
                chart_row += 2
                for text_file in txt_query_list:
                    title = text_file.replace("_", " ").upper()
                    # print(device_name)
                    results = self.generate_report(
                        text_file, from_date, to_date, organizations,
                        device_name)
                    if len(results) > 1:
                        we = wb.create_sheet("data" + str(min_row), 1)
                        we.sheet_state = 'hidden'
                        self.generate_chart(
                            ws, we, title, results, chart_row, min_row, from_date, to_date)
                        chart_row += 22
                        min_row += len(results)
                    else:
                        chart_row += 2
                        query_to_write = "No " + str(
                            title) + " Chart Found for device – " \
                            + device_name + " for selected time range"
                        self.write_data_on_sheet(ws, chart_row, query_to_write)
        else:
            ws['A' + str(chart_row)] = "No Data Found for selected device – "
            ws['A' + str(chart_row)].font = fontStyle
        import datetime
        x = datetime.datetime.now()
        date_time_str = x.strftime("%d_%m_%YT%H_%M_%S")
        file_name = "xls_data_" + utils.gen_random() + str(date_time_str + ".xlsx")
        wb.save(str(utils.get_report_path()) + file_name)
        return file_name

    def generate_report(
        self, textfile, from_date, to_date, organizations, devices_by_organization
    ):
        query_list = utils.frame_query(str(textfile))
        data_event = {}
        data_date = {}

        for query in query_list.split("===next==="):
            params = (devices_by_organization, from_date, to_date,)
            rows = self.sciencelogic_db.execute(query, params, True)
            rows.pop(0)
            for res in rows:
                device_name = res[2]
                device_date = res[0]
                if data_event.get(device_name) is None:
                    temp_list = [res]
                    data_event[device_name] = temp_list
                else:
                    data_event.get(device_name).append(res)

                if data_date.get(device_date) is None:
                    temp_list = [res]
                    data_date[device_date] = temp_list
                else:
                    data_date.get(device_date).append(res)
        final_list = []
        for key, value in data_date.items():
            temp_list = [key]
            for dt, vl in data_event.items():
                flag = True
                for i in vl:
                    if key == i[0]:
                        flag = False
                        temp_list.append(float(i[1]))
                if flag:
                    temp_list.append(float('0'))
            final_list.append(temp_list)
        headers = ['dates']
        headers.extend(data_event.keys())
        final_list.insert(0, headers)
        return final_list

    def generate_chart(
        self, ws, we, title, results, chart_row, min_row, from_date, to_date
    ):
        from_flag = True
        to_flag = True
        count = 0
        for row in results:
            if from_date in row:
                from_flag = False
            if to_date in row:
                to_flag = False
            we.append(row)
        if from_flag:
            we.insert_rows(2, amount=1)
            we.cell(2, 1, from_date)
            count += 1
        if to_flag:
            we.append([str(to_date)])
            count += 1

        if len(results) > 1:
            c2 = LineChart()
            c2.title = title
            # c2.style = 12
            c2.y_axis.title = "Packet loss count"
            c2.y_axis.crossAx = 500
            c2.height = 10  # default is 7.5
            # width = len(result) - 1
            # width = width if width > 15 else 15
            # width = 30
            c2.width = 30
            c2.x_axis = DateAxis(crossAx=100)
            c2.x_axis.number_format = 'd-mmm-yyyy HH:MM:SS'
            c2.x_axis.majorTimeUnit = "days"
            c2.x_axis.title = "Date"
            max_col = len(results[0])

            data = Reference(
                we, min_col=2, min_row=1,
                max_col=max_col, max_row=len(results) + count)
            c2.add_data(data, titles_from_data=True)
            dates = Reference(
                we, min_col=1, min_row=2,
                max_row=len(results) + count)
            c2.set_categories(dates)

            flag = True
            if len(results) == 2:
                flag = False
                for c in c2.series:
                    # 12700 emu = 1 pt(72*1 inches)
                    c.graphicalProperties.line.width = 20000
                    c.marker.symbol = 'diamond'
            if flag:
                for c in c2.series:
                    # 12700 emu = 1 pt(72*1 inches)
                    c.graphicalProperties.line.width = 20000

            ws.add_chart(c2, "A" + str(chart_row))

    def write_data_on_sheet(self, ws, chart_row, data):
        font = Font(bold=True, size="12")
        ws['A' + str(chart_row)] = data
        ws['A' + str(chart_row)].font = font
        ws['A' + str(chart_row)].fill = PatternFill(
            start_color="00999999",
            end_color="999999", fill_type="solid")
        ws.merge_cells('A' + str(chart_row) + ':F' + str(chart_row))
